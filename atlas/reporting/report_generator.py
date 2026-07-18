"""
Report Generator

Converts the output of the Ranking Engine into a standardized reporting format.
"""
import logging
from pathlib import Path
from typing import Dict, Any
from datetime import datetime

import pandas as pd

logger = logging.getLogger(__name__)


class ReportError(Exception):
    """Raised when report generation validation fails."""
    pass


class ReportGenerator:
    """
    Generates presentation-ready reports from ranking engine dataframes.
    """

    REQUIRED_COLUMNS = ["Rank", "Symbol", "Score"]

    def __init__(self) -> None:
        """Initialize the ReportGenerator."""
        logger.debug("Initialized ReportGenerator")

    def generate(self, ranking_dataframe: pd.DataFrame, universe_name: str = "Unknown") -> Dict[str, Any]:
        """
        Generate a summary report from the ranking dataframe.

        Args:
            ranking_dataframe (pd.DataFrame): The output from the RankingEngine.
            universe_name (str): The name of the scanned universe. Defaults to "Unknown".

        Returns:
            Dict[str, Any]: A dictionary containing summary statistics and the dataframe.
            
        Raises:
            ReportError: If the dataframe is empty or missing required columns.
        """
        if ranking_dataframe is None or ranking_dataframe.empty:
            logger.error("Report generation failed: dataframe is empty")
            raise ReportError("Input dataframe cannot be empty.")
            
        missing = [col for col in self.REQUIRED_COLUMNS if col not in ranking_dataframe.columns]
        if missing:
            logger.error(f"Missing required columns for ReportGenerator: {missing}")
            raise ReportError(f"Missing required columns: {missing}")

        logger.info(f"Generating report for universe: {universe_name}")
        
        # We assume all rows in the dataframe were scanned.
        # Qualified stocks are those with a Score > 0.
        stocks_scanned = len(ranking_dataframe)
        qualified_df = ranking_dataframe[ranking_dataframe["Score"] > 0]
        qualified = len(qualified_df)
        
        top_score = 0
        average_score = 0.0
        
        if qualified > 0:
            top_score = float(qualified_df["Score"].max())
            average_score = float(qualified_df["Score"].mean())
            
        # Use current date
        current_date = datetime.now().strftime("%Y-%m-%d")
        
        report = {
            "date": current_date,
            "universe": universe_name,
            "stocks_scanned": stocks_scanned,
            "qualified": qualified,
            "top_score": top_score,
            "average_score": round(average_score, 2),
            "results": ranking_dataframe.copy()
        }
        
        return report

    def to_text(self, report: Dict[str, Any]) -> str:
        """
        Format the report dictionary into a readable text string.

        Args:
            report (Dict[str, Any]): The generated report dictionary.

        Returns:
            str: The formatted text report.
        """
        lines = []
        lines.append("====================================")
        lines.append("MULTIBAGGER DAILY REPORT")
        lines.append("====================================")
        lines.append(f"Date: {report.get('date', '')}")
        lines.append(f"Universe: {report.get('universe', '')}")
        lines.append(f"Stocks Scanned: {report.get('stocks_scanned', 0)}")
        lines.append(f"Qualified: {report.get('qualified', 0)}")
        lines.append(f"Top Score: {report.get('top_score', 0)}")
        lines.append(f"Average Score: {report.get('average_score', 0.0)}")
        lines.append("------------------------------------")
        
        df = report.get("results")
        if df is not None and not df.empty:
            # We want to show Rank, Symbol, Score as per the requested formatting
            cols = [c for c in ["Rank", "Symbol", "Score"] if c in df.columns]
            df_string = df[cols].to_string(index=False)
            lines.append(df_string)
        else:
            lines.append("No results.")
            
        lines.append("------------------------------------")
        
        return "\n".join(lines)

    def export_csv(self, report: Dict[str, Any], output_path: Path) -> Path:
        """
        Export the ranked stocks to a CSV file.
        
        Args:
            report (Dict[str, Any]): The generated report dictionary.
            output_path (Path): The path where the CSV should be saved.
            
        Returns:
            Path: The path to the saved CSV file.
            
        Raises:
            ReportError: If exporting fails.
        """
        df = report.get("results")
        if df is None:
            raise ReportError("Report dictionary is missing 'results' dataframe.")
            
        try:
            # Ensure the directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Requested exact column order
            columns = ["Rank", "Symbol", "Score", "Institutional", "Momentum", "VCP"]
            
            # Start with an empty dataframe that enforces the column order
            export_df = pd.DataFrame(columns=columns)
            
            if not df.empty:
                temp_df = df.copy()
                
                # Standardize column casing
                rename_map = {
                    "symbol": "Symbol",
                    "institutional": "Institutional",
                    "momentum": "Momentum",
                    "vcp": "VCP",
                    "rank": "Rank",
                    "score": "Score"
                }
                temp_df = temp_df.rename(columns=rename_map)
                
                # Find which of the requested columns are actually available
                available_cols = [c for c in columns if c in temp_df.columns]
                
                # Concatenate vertically
                if available_cols:
                    export_df = pd.concat([export_df, temp_df[available_cols]], ignore_index=True)
            
            # Save using utf-8 encoding as required
            export_df[columns].to_csv(output_path, index=False, encoding="utf-8")
            logger.info(f"Report exported to {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to export CSV report: {e}")
            raise ReportError(f"Failed to export CSV report: {e}")

    def export_excel(self, report: Dict[str, Any], output_path: Path) -> Path:
        """
        Export the ranked stocks to a formatted Excel workbook.
        
        Args:
            report (Dict[str, Any]): The generated report dictionary.
            output_path (Path): The path where the Excel file should be saved.
            
        Returns:
            Path: The path to the saved Excel file.
            
        Raises:
            ReportError: If exporting fails.
        """
        df = report.get("results")
        if df is None:
            raise ReportError("Report dictionary is missing 'results' dataframe.")
            
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            columns = ["Rank", "Symbol", "Score", "Institutional", "Momentum", "VCP"]
            export_df = pd.DataFrame(columns=columns)
            
            if not df.empty:
                temp_df = df.copy()
                rename_map = {
                    "symbol": "Symbol",
                    "institutional": "Institutional",
                    "momentum": "Momentum",
                    "vcp": "VCP",
                    "rank": "Rank",
                    "score": "Score"
                }
                temp_df = temp_df.rename(columns=rename_map)
                available_cols = [c for c in columns if c in temp_df.columns]
                if available_cols:
                    export_df = pd.concat([export_df, temp_df[available_cols]], ignore_index=True)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                export_df[columns].to_excel(writer, sheet_name='Daily Scan', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Daily Scan']
                
                from openpyxl.styles import Font, Alignment, PatternFill
                from openpyxl.formatting.rule import CellIsRule

                # 5. Freeze the first row
                worksheet.freeze_panes = "A2"
                
                # 6. Enable AutoFilter
                last_col_letter = chr(ord('A') + len(columns) - 1)
                worksheet.auto_filter.ref = f"A1:{last_col_letter}{max(1, len(export_df)+1)}"
                
                # Formatting objects
                bold_font = Font(bold=True)
                center_align = Alignment(horizontal='center', vertical='center')
                
                for col_idx, col_name in enumerate(columns, start=1):
                    # 7. Bold header
                    header_cell = worksheet.cell(row=1, column=col_idx)
                    header_cell.font = bold_font
                    
                    if col_name != "Symbol":
                        header_cell.alignment = center_align
                
                # Column adjustments and data alignment
                for col_idx, col_name in enumerate(columns, start=1):
                    col_letter = chr(ord('A') + col_idx - 1)
                    
                    # 4. Auto-adjust column widths
                    max_length = len(col_name)
                    if not export_df.empty:
                        col_values = export_df[col_name].astype(str)
                        max_length = max(max_length, col_values.map(len).max())
                    
                    worksheet.column_dimensions[col_letter].width = max_length + 2
                    
                    # 8. Center align every column except Symbol
                    if col_name != "Symbol":
                        for row_idx in range(2, len(export_df) + 2):
                            worksheet.cell(row=row_idx, column=col_idx).alignment = center_align
                
                # 9. Apply conditional formatting to Score column (C)
                if not export_df.empty:
                    score_range = f"C2:C{len(export_df)+1}"
                    
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                    
                    worksheet.conditional_formatting.add(
                        score_range,
                        CellIsRule(operator='greaterThanOrEqual', formula=['70'], fill=green_fill)
                    )
                    worksheet.conditional_formatting.add(
                        score_range,
                        CellIsRule(operator='between', formula=['40', '69'], fill=yellow_fill)
                    )
                    worksheet.conditional_formatting.add(
                        score_range,
                        CellIsRule(operator='between', formula=['1', '39'], fill=blue_fill)
                    )
                    
            logger.info(f"Report exported to Excel: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to export Excel report: {e}")
            raise ReportError(f"Failed to export Excel report: {e}")

    def append_history(self, report: Dict[str, Any], history_path: Path) -> Path:
        """
        Append the daily scan results to a historical archive CSV.
        
        Args:
            report (Dict[str, Any]): The generated report dictionary.
            history_path (Path): The path to the master history CSV.
            
        Returns:
            Path: The path to the history file.
        """
        df = report.get("results")
        if df is None:
            raise ReportError("Report dictionary is missing 'results' dataframe.")
            
        try:
            history_path.parent.mkdir(parents=True, exist_ok=True)
            
            columns = ["ScanDate", "Rank", "Symbol", "Score", "Institutional", "Momentum", "VCP"]
            new_data = pd.DataFrame(columns=columns)
            
            if not df.empty:
                temp_df = df.copy()
                rename_map = {
                    "symbol": "Symbol",
                    "institutional": "Institutional",
                    "momentum": "Momentum",
                    "vcp": "VCP",
                    "rank": "Rank",
                    "score": "Score"
                }
                temp_df = temp_df.rename(columns=rename_map)
                temp_df["ScanDate"] = report.get("date", datetime.now().strftime("%Y-%m-%d"))
                
                available_cols = [c for c in columns if c in temp_df.columns]
                new_data = pd.concat([new_data, temp_df[available_cols]], ignore_index=True)
            
            if history_path.exists():
                history_df = pd.read_csv(history_path)
                # Ensure existing history has the required columns
                for col in columns:
                    if col not in history_df.columns:
                        history_df[col] = None
                        
                combined_df = pd.concat([history_df, new_data], ignore_index=True)
                
                # Drop duplicates for the same ScanDate and Symbol
                # This ensures we don't duplicate stocks if we run multiple scans in one day
                combined_df = combined_df.drop_duplicates(subset=["ScanDate", "Symbol"], keep="last")
                
                # Preserve chronological order
                combined_df = combined_df.sort_values(by=["ScanDate", "Rank"])
            else:
                combined_df = new_data
                if not combined_df.empty:
                    combined_df = combined_df.sort_values(by=["ScanDate", "Rank"])
            
            # Export back to CSV
            combined_df[columns].to_csv(history_path, index=False, encoding="utf-8")
            logger.info(f"Scan history appended to {history_path}")
            return history_path
            
        except Exception as e:
            logger.error(f"Failed to append to scan history: {e}")
            raise ReportError(f"Failed to append to scan history: {e}")

    def export_html(self, report: Dict[str, Any], output_path: Path) -> Path:
        """
        Export the daily scan results to a standalone HTML file.
        
        Args:
            report (Dict[str, Any]): The generated report dictionary.
            output_path (Path): The path where the HTML file should be saved.
            
        Returns:
            Path: The path to the saved HTML file.
            
        Raises:
            ReportError: If exporting fails.
        """
        df = report.get("results")
        if df is None:
            raise ReportError("Report dictionary is missing 'results' dataframe.")
            
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            html_content = [
                "<!DOCTYPE html>",
                "<html lang='en'>",
                "<head>",
                "<meta charset='UTF-8'>",
                "<meta name='viewport' content='width=device-width, initial-scale=1.0'>",
                "<title>Atlas Daily Scan Report</title>",
                "<style>",
                "body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f8f9fa; color: #333; margin: 0; padding: 20px; }",
                ".container { max-width: 1400px; margin: 0 auto; background: #fff; padding: 20px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }",
                "h1 { color: #2c3e50; text-align: center; border-bottom: 2px solid #eee; padding-bottom: 10px; }",
                ".summary-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 20px 0; }",
                ".stat-box { background: #f1f3f5; padding: 15px; border-radius: 6px; text-align: center; }",
                ".stat-title { font-size: 0.9em; color: #6c757d; text-transform: uppercase; font-weight: bold; }",
                ".stat-value { font-size: 1.5em; font-weight: bold; color: #212529; margin-top: 5px; }",
                ".table-container { overflow-x: auto; margin-top: 20px; max-height: 70vh; }",
                "table { width: 100%; border-collapse: collapse; text-align: center; white-space: nowrap; }",
                "th { position: sticky; top: 0; background-color: #2c3e50; color: #fff; padding: 12px; font-weight: 600; z-index: 1; }",
                "td { padding: 10px; border-bottom: 1px solid #eee; }",
                "tr:nth-child(even) { background-color: #f8f9fa; }",
                "tr:hover { background-color: #e9ecef; }",
                ".badge { padding: 4px 8px; border-radius: 12px; font-size: 0.85em; font-weight: bold; color: #fff; }",
                ".badge-true { background-color: #28a745; }",
                ".badge-false { background-color: #6c757d; }",
                ".score-green { background-color: #d4edda; color: #155724; font-weight: bold; }",
                ".score-yellow { background-color: #fff3cd; color: #856404; font-weight: bold; }",
                ".score-blue { background-color: #cce5ff; color: #004085; font-weight: bold; }",
                ".score-white { background-color: #fff; color: #333; }",
                ".empty-msg { text-align: center; padding: 30px; font-size: 1.2em; color: #6c757d; font-style: italic; }",
                "</style>",
                "</head>",
                "<body>",
                "<div class='container'>",
                "<h1>Atlas Daily Scan Report</h1>",
                "<div class='summary-grid'>"
            ]
            
            stats = [
                ("Scan Date", report.get("date", "")),
                ("Universe", report.get("universe", "")),
                ("Stocks Scanned", report.get("stocks_scanned", 0)),
                ("Qualified Stocks", report.get("qualified", 0)),
                ("Top Score", report.get("top_score", 0)),
                ("Average Score", report.get("average_score", 0.0))
            ]
            
            for title, value in stats:
                html_content.append(f"<div class='stat-box'><div class='stat-title'>{title}</div><div class='stat-value'>{value}</div></div>")
            
            html_content.append("</div>")
            
            if df.empty or report.get("qualified", 0) == 0:
                html_content.append("<div class='empty-msg'>No qualifying stocks today.</div>")
            else:
                html_content.append("<div class='table-container'>")
                html_content.append("<table>")
                
                target_cols = [
                    "Rank", "Symbol", "Score", "Institutional", "Momentum", "VCP", 
                    "MomentumScore", "RVOL20", "RS90", "EMA20", "EMA50", "EMA200", "Close"
                ]
                
                temp_df = df.copy()
                rename_map = {
                    "rank": "Rank", "symbol": "Symbol", "score": "Score", 
                    "institutional": "Institutional", "momentum": "Momentum", "vcp": "VCP"
                }
                temp_df = temp_df.rename(columns=rename_map)
                
                html_content.append("<tr>")
                for col in target_cols:
                    html_content.append(f"<th>{col}</th>")
                html_content.append("</tr>")
                
                # Ensure Score column exists and is numeric to filter properly
                if "Score" in temp_df.columns:
                    temp_df["Score"] = pd.to_numeric(temp_df["Score"], errors='coerce').fillna(0)
                    qualified_df = temp_df[temp_df["Score"] > 0]
                else:
                    qualified_df = pd.DataFrame(columns=target_cols)
                
                for _, row in qualified_df.iterrows():
                    html_content.append("<tr>")
                    for col in target_cols:
                        val = row.get(col, "")
                        
                        if col == "Score":
                            try:
                                score_val = float(val) if val != "" else 0
                                if score_val >= 70:
                                    css_class = "score-green"
                                elif score_val >= 40:
                                    css_class = "score-yellow"
                                elif score_val >= 1:
                                    css_class = "score-blue"
                                else:
                                    css_class = "score-white"
                            except ValueError:
                                css_class = "score-white"
                                
                            html_content.append(f"<td class='{css_class}'>{val}</td>")
                            
                        elif col in ["Institutional", "Momentum", "VCP"]:
                            bool_val = bool(val) if val != "" else False
                            if bool_val:
                                html_content.append("<td><span class='badge badge-true'>True</span></td>")
                            else:
                                html_content.append("<td><span class='badge badge-false'>False</span></td>")
                                
                        elif isinstance(val, float):
                            html_content.append(f"<td>{val:.2f}</td>")
                            
                        else:
                            html_content.append(f"<td>{val}</td>")
                            
                    html_content.append("</tr>")
                    
                html_content.append("</table>")
                html_content.append("</div>")
                
            html_content.append("</div>")
            html_content.append("</body>")
            html_content.append("</html>")
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("\n".join(html_content))
                
            logger.info(f"Report exported to HTML: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to export HTML report: {e}")
            raise ReportError(f"Failed to export HTML report: {e}")
