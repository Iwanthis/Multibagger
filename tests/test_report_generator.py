import pandas as pd
import pytest
from datetime import datetime

from atlas.reporting.report_generator import ReportGenerator, ReportError


@pytest.fixture
def sample_ranking_df():
    return pd.DataFrame({
        "Rank": [1, 2, 3, 4],
        "Symbol": ["CDSL", "HAL", "RELIANCE", "HDFCBANK"],
        "Score": [96, 92, 90, 0],
        "Institutional": [True, True, True, False],
        "Momentum": [True, True, False, False]
    })


def test_report_generation(sample_ranking_df):
    generator = ReportGenerator()
    report = generator.generate(sample_ranking_df, universe_name="Nifty500")
    
    assert report["universe"] == "Nifty500"
    assert report["stocks_scanned"] == 4
    assert report["qualified"] == 3 # Only > 0 scores
    assert report["top_score"] == 96
    # Average of 96, 92, 90 = 278 / 3 = 92.67
    assert report["average_score"] == 92.67
    assert report["date"] == datetime.now().strftime("%Y-%m-%d")
    
    # Assert dataframe is passed through correctly
    assert isinstance(report["results"], pd.DataFrame)
    assert len(report["results"]) == 4


def test_text_formatting(sample_ranking_df):
    generator = ReportGenerator()
    report = generator.generate(sample_ranking_df, universe_name="Nifty500")
    
    text = generator.to_text(report)
    
    assert "MULTIBAGGER DAILY REPORT" in text
    assert "Universe: Nifty500" in text
    assert "Stocks Scanned: 4" in text
    assert "Qualified: 3" in text
    assert "Top Score: 96" in text
    assert "Average Score: 92.67" in text
    assert "CDSL" in text
    assert "Rank" in text


def test_empty_dataframe():
    df = pd.DataFrame(columns=["Rank", "Symbol", "Score"])
    generator = ReportGenerator()
    
    with pytest.raises(ReportError, match="cannot be empty"):
        generator.generate(df)


def test_missing_columns():
    df = pd.DataFrame({
        "Symbol": ["A", "B"],
        "Score": [10, 20]
        # Missing 'Rank'
    })
    generator = ReportGenerator()
    
    with pytest.raises(ReportError, match="Missing required columns"):
        generator.generate(df)


def test_no_qualified_stocks():
    df = pd.DataFrame({
        "Rank": [1, 2],
        "Symbol": ["A", "B"],
        "Score": [0, 0]
    })
    generator = ReportGenerator()
    report = generator.generate(df)
    
    assert report["qualified"] == 0
    assert report["top_score"] == 0
    assert report["average_score"] == 0.0


def test_dataframe_unchanged(sample_ranking_df):
    original = sample_ranking_df.copy(deep=True)
    generator = ReportGenerator()
    generator.generate(sample_ranking_df)
    
    pd.testing.assert_frame_equal(sample_ranking_df, original)


def test_export_csv_success(sample_ranking_df, tmp_path):
    generator = ReportGenerator()
    report = generator.generate(sample_ranking_df)
    
    export_path = tmp_path / "reports" / "test_export.csv"
    returned_path = generator.export_csv(report, export_path)
    
    assert returned_path == export_path
    assert export_path.exists()
    
    # Verify content
    df = pd.read_csv(export_path)
    assert list(df.columns) == ["Rank", "Symbol", "Score", "Institutional", "Momentum", "VCP"]
    assert len(df) == 4
    assert df.iloc[0]["Symbol"] == "CDSL"
    # Ensure booleans were preserved
    assert df.iloc[0]["Institutional"] == True


def test_export_csv_empty_report(tmp_path):
    # Tests that an empty scan still generates a valid CSV with just headers
    empty_df = pd.DataFrame({
        "Rank": [], "Symbol": [], "Score": [], 
        "Institutional": [], "Momentum": [], "VCP": []
    })
    
    generator = ReportGenerator()
    # Mock the report dict directly since generate() fails on empty DF
    report = {"results": empty_df}
    
    export_path = tmp_path / "empty_export.csv"
    generator.export_csv(report, export_path)
    
    assert export_path.exists()
    df = pd.read_csv(export_path)
    assert len(df) == 0
    assert list(df.columns) == ["Rank", "Symbol", "Score", "Institutional", "Momentum", "VCP"]


def test_export_csv_missing_results(tmp_path):
    generator = ReportGenerator()
    report = {} # Missing 'results'
    export_path = tmp_path / "fail.csv"
    
    with pytest.raises(ReportError, match="missing 'results'"):
        generator.export_csv(report, export_path)


def test_export_csv_directory_creation(sample_ranking_df, tmp_path):
    generator = ReportGenerator()
    report = generator.generate(sample_ranking_df)
    
    # Deep nested path
    export_path = tmp_path / "deep" / "nested" / "dir" / "report.csv"
    assert not export_path.parent.exists()
    
    generator.export_csv(report, export_path)
    
    assert export_path.parent.exists()
    assert export_path.exists()

def test_export_excel_success(sample_ranking_df, tmp_path):
    generator = ReportGenerator()
    report = generator.generate(sample_ranking_df)
    
    export_path = tmp_path / "test_export.xlsx"
    returned_path = generator.export_excel(report, export_path)
    
    assert returned_path == export_path
    assert export_path.exists()
    
    # Read the excel back to verify content
    df = pd.read_excel(export_path, engine='openpyxl')
    assert list(df.columns) == ["Rank", "Symbol", "Score", "Institutional", "Momentum", "VCP"]
    assert len(df) == 4
    assert df.iloc[0]["Symbol"] == "CDSL"
    
    # Verify openpyxl formatting (worksheet name, freeze panes, autofilter)
    import openpyxl
    wb = openpyxl.load_workbook(export_path)
    assert "Daily Scan" in wb.sheetnames
    ws = wb["Daily Scan"]
    
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith("A1:F")

def test_export_excel_empty_report(tmp_path):
    empty_df = pd.DataFrame({
        "Rank": [], "Symbol": [], "Score": [], 
        "Institutional": [], "Momentum": [], "VCP": []
    })
    generator = ReportGenerator()
    report = {"results": empty_df}
    
    export_path = tmp_path / "empty_export.xlsx"
    generator.export_excel(report, export_path)
    
    assert export_path.exists()
    import openpyxl
    wb = openpyxl.load_workbook(export_path)
    ws = wb["Daily Scan"]
    assert ws.max_row == 1 # Only header
    
    df = pd.read_excel(export_path, engine='openpyxl')
    assert len(df) == 0

def test_export_excel_missing_results(tmp_path):
    generator = ReportGenerator()
    report = {}
    export_path = tmp_path / "fail.xlsx"
    
    with pytest.raises(ReportError, match="missing 'results'"):
        generator.export_excel(report, export_path)
    
def test_append_history_creation(sample_ranking_df, tmp_path):
    generator = ReportGenerator()
    report = generator.generate(sample_ranking_df)
    
    history_path = tmp_path / "scan_history.csv"
    assert not history_path.exists()
    
    returned_path = generator.append_history(report, history_path)
    
    assert returned_path == history_path
    assert history_path.exists()
    
    df = pd.read_csv(history_path)
    assert len(df) == 4
    assert list(df.columns) == ["ScanDate", "Rank", "Symbol", "Score", "Institutional", "Momentum", "VCP"]
    assert df.iloc[0]["Symbol"] == "CDSL"
    assert df.iloc[0]["ScanDate"] == report["date"]


def test_append_history_duplicate_prevention_and_append(sample_ranking_df, tmp_path):
    generator = ReportGenerator()
    report = generator.generate(sample_ranking_df)
    
    history_path = tmp_path / "scan_history.csv"
    
    # Run 1: First append
    generator.append_history(report, history_path)
    
    # Run 2: Exact same day (should overwrite/deduplicate)
    # CDSL score changes from 96 to 99 in memory
    report["results"].loc[report["results"]["Symbol"] == "CDSL", "Score"] = 99
    generator.append_history(report, history_path)
    
    df_after_run2 = pd.read_csv(history_path)
    # Length should still be 4 because it dropped duplicates for same ScanDate + Symbol
    assert len(df_after_run2) == 4
    assert df_after_run2[df_after_run2["Symbol"] == "CDSL"].iloc[0]["Score"] == 99
    
    # Run 3: Next day
    report["date"] = "2099-01-01"
    generator.append_history(report, history_path)
    
    df_after_run3 = pd.read_csv(history_path)
    # Length should now be 8 (4 from today, 4 from 2099)
    assert len(df_after_run3) == 8


def test_append_history_chronological_ordering(sample_ranking_df, tmp_path):
    generator = ReportGenerator()
    history_path = tmp_path / "scan_history.csv"
    
    report1 = generator.generate(sample_ranking_df)
    report1["date"] = "2023-01-02"
    generator.append_history(report1, history_path)
    
    # Append older data (should sort correctly)
    report2 = generator.generate(sample_ranking_df)
    report2["date"] = "2023-01-01"
    generator.append_history(report2, history_path)
    
    df = pd.read_csv(history_path)
    assert len(df) == 8
    # 2023-01-01 should appear before 2023-01-02
    assert df.iloc[0]["ScanDate"] == "2023-01-01"
    assert df.iloc[4]["ScanDate"] == "2023-01-02"


def test_append_history_dataframe_unchanged(sample_ranking_df, tmp_path):
    original = sample_ranking_df.copy(deep=True)
    generator = ReportGenerator()
    report = generator.generate(sample_ranking_df)
    history_path = tmp_path / "scan_history.csv"
    
    generator.append_history(report, history_path)
    pd.testing.assert_frame_equal(sample_ranking_df, original)
